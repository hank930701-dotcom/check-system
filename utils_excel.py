from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import os

EXCEL_FILE = "output/report.xlsx"


def save_to_excel(data, images):

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"

        headers = [
            "日期","人員","預估Sales","實際Sales",
            "矩陣工時","排班工時","實際工時","訓練工時",
            "工時超用說明","超8","TPLH","SPMH","會員數",
            "汽水機噴嘴","汽水機座閥","盤點差異表","盤點過帳",
            "損耗過帳","K區地面","裹粉台","炸鍋",
            "炸鍋底部","二樓地面","三樓地面","垃圾桶","廁所"
        ]

        ws.append(headers)

    row = [
        data.get("date"),
        data.get("person"),
        data.get("forecast_sales"),
        data.get("actual_sales"),
        data.get("matrix_hours"),
        data.get("schedule_hours"),
        data.get("actual_hours"),
        data.get("training_hours"),
        data.get("note"),
        data.get("over8"),
        data.get("tplh"),
        data.get("spmh"),
        data.get("member")
    ]

    ws.append(row)
    row_idx = ws.max_row

    img_fields = [
        "soda_nozzle","soda_valve","inventory_diff","posting_screen",
        "loss_screen","k_floor","batter_table","fryer",
        "fryer_bottom","floor2","floor3","trash","toilet"
    ]

    start_col = 14

    for i, field in enumerate(img_fields):
        if field in images:
            img = Image(images[field])
            img.width = 120
            img.height = 120

            col = start_col + i
            cell = ws.cell(row=row_idx, column=col)
            ws.add_image(img, cell.coordinate)

    wb.save(EXCEL_FILE)
